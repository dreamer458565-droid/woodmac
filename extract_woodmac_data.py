#!/usr/bin/env python3
"""
Wood Mackenzie Nickel Market Data Extraction Script
Extracts data from STO (Monthly) and IHO (Quarterly) Excel files
and produces a comprehensive JSON dashboard dataset.
"""

import json
import xlrd
from openpyxl import load_workbook
from datetime import datetime
import math

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "데이터"

def clean_value(val):
    """Convert Excel values to clean Python types"""
    if val is None or val == '':
        return None
    if isinstance(val, str) and val.strip() == '':
        return None
    if isinstance(val, (int, float)):
        if math.isnan(val):
            return None
        return val
    try:
        # Try to convert string to float
        return float(val)
    except (ValueError, TypeError):
        # Keep as string if not numeric
        return str(val) if val else None

def safe_divide(num, denom):
    """Safely divide two values"""
    if num is None or denom is None or denom == 0:
        return None
    return num / denom

# File paths
STO_FILE = str(DATA_DIR / "nickel_sto_data_tables_february_2026.xls")
IHO_DEMAND_FILE = str(DATA_DIR / "global_nickel_investment_horizon_outlook_q4_2025/nickel-demand-analysis-december-2025.xlsx")
IHO_BALANCE_FILE = str(DATA_DIR / "global_nickel_investment_horizon_outlook_q4_2025/nickel-market-balance-prices-december-2025.xls")
IHO_REFINERY_FILE = str(DATA_DIR / "global_nickel_investment_horizon_outlook_q4_2025/nickel-refinery-analysis-december-2025.xls")

data = {
    "metadata": {
        "extraction_date": datetime.now().isoformat(),
        "source_files": {
            "sto_monthly": "nickel_sto_data_tables_february_2026.xls",
            "iho_demand": "nickel-demand-analysis-december-2025.xlsx",
            "iho_balance": "nickel-market-balance-prices-december-2025.xls",
            "iho_refinery": "nickel-refinery-analysis-december-2025.xls"
        },
        "version": "1.0"
    },
    "tab_1_global_market_balance": {},
    "tab_2_ev_battery_nickel": {},
    "tab_3_ambatovy": {},
    "tab_4_south_korea": {},
    "tab_5_supply": {}
}

# ============================================================================
# TAB 1: GLOBAL MARKET BALANCE
# ============================================================================
print("Extracting TAB 1: Global Market Balance...")

# IHO Market Balance - Imbalance sheet
iho_balance_wb = xlrd.open_workbook(IHO_BALANCE_FILE)
imbalance_ws = iho_balance_wb.sheet_by_name('Imbalance')

# Extract years from row 4 (column 2 onwards)
years = []
imbalance_data = {
    "production_capability": {},
    "refined_output": {},
    "secondary_ev_recycling": {},
    "primary_supply": {},
    "consumption": {},
    "surplus_deficit": {}
}

for col in range(2, min(imbalance_ws.ncols, 40)):
    year = imbalance_ws.cell_value(4, col)
    if year and isinstance(year, (int, float)):
        year = int(year)
        years.append(year)
        # Row 6: Production Capability
        imbalance_data["production_capability"][year] = clean_value(imbalance_ws.cell_value(6, col))
        # Row 12: Refined Output
        imbalance_data["refined_output"][year] = clean_value(imbalance_ws.cell_value(12, col))
        # Row 19: Secondary/EV recycling
        imbalance_data["secondary_ev_recycling"][year] = clean_value(imbalance_ws.cell_value(19, col))
        # Row 26: Primary Supply
        imbalance_data["primary_supply"][year] = clean_value(imbalance_ws.cell_value(26, col))
        # Row 28: Consumption
        imbalance_data["consumption"][year] = clean_value(imbalance_ws.cell_value(28, col))
        # Row 30: Surplus/Deficit
        imbalance_data["surplus_deficit"][year] = clean_value(imbalance_ws.cell_value(30, col))

data["tab_1_global_market_balance"]["iho_imbalance"] = imbalance_data

# STO QuarterlyBalance sheet
sto_wb = xlrd.open_workbook(STO_FILE)
qb_ws = sto_wb.sheet_by_name('QuarterlyBalance')

quarterly_balance = {}
# Row 9 onwards has data; Year in col 0, Quarter in col 1
for row in range(9, min(qb_ws.nrows, 33)):
    year_val = qb_ws.cell_value(row, 0)
    quarter = qb_ws.cell_value(row, 1)
    if year_val and isinstance(year_val, (int, float)):
        year = int(year_val)
        if year not in quarterly_balance:
            quarterly_balance[year] = {}

        quarter_key = f"Q{int(quarter[1])}" if isinstance(quarter, str) and quarter.startswith('Q') else str(quarter)

        quarterly_balance[year][quarter_key] = {
            "refined_supply_kt": clean_value(qb_ws.cell_value(row, 2)),
            "refined_consumption_kt": clean_value(qb_ws.cell_value(row, 3)),
            "balance_kt": clean_value(qb_ws.cell_value(row, 4)),
            "lme_price_usd_tonne": clean_value(qb_ws.cell_value(row, 6)),
            "lme_price_c_lb": clean_value(qb_ws.cell_value(row, 7)),
            "stocks_kt": clean_value(qb_ws.cell_value(row, 9))
        }

data["tab_1_global_market_balance"]["sto_quarterly_balance"] = quarterly_balance

# STO BalancesByClass sheet
bc_ws = sto_wb.sheet_by_name('BalancesByClass')
balances_by_class = {}
# Row 6 has years in columns 1-5
years_bc = []
for col in range(1, 6):
    year = bc_ws.cell_value(6, col)
    if year:
        years_bc.append(int(year))

balances_by_class = {
    "class_i": {},
    "class_ii": {},
    "sulphate": {},
    "total_nickel": {}
}

for idx, year in enumerate(years_bc, start=1):
    # Class I - rows 8-10
    balances_by_class["class_i"][year] = {
        "china": clean_value(bc_ws.cell_value(8, idx)),
        "rest_of_world": clean_value(bc_ws.cell_value(9, idx)),
        "total": clean_value(bc_ws.cell_value(10, idx))
    }
    # Class II - rows 13-15
    balances_by_class["class_ii"][year] = {
        "china": clean_value(bc_ws.cell_value(13, idx)),
        "rest_of_world": clean_value(bc_ws.cell_value(14, idx)),
        "total": clean_value(bc_ws.cell_value(15, idx))
    }
    # Sulphate - rows 18-20
    balances_by_class["sulphate"][year] = {
        "china": clean_value(bc_ws.cell_value(18, idx)),
        "rest_of_world": clean_value(bc_ws.cell_value(19, idx)),
        "total": clean_value(bc_ws.cell_value(20, idx))
    }
    # Total - row 22
    balances_by_class["total_nickel"][year] = clean_value(bc_ws.cell_value(22, idx))

data["tab_1_global_market_balance"]["sto_balances_by_class"] = balances_by_class

# ============================================================================
# TAB 2: EV BATTERY NICKEL DEMAND
# ============================================================================
print("Extracting TAB 2: EV Battery Nickel Demand...")

# STO NiInPrecursors sheet
nip_ws = sto_wb.sheet_by_name('NiInPrecursors')

# Years are in row 6, columns 1-5
years_nip = []
for col in range(1, 6):
    year = nip_ws.cell_value(6, col)
    if year:
        years_nip.append(int(year))

battery_demand_by_country = {}
# Rows 8-26 are country data
country_rows = {
    8: "Other Africa",
    9: "Total Africa",
    10: "China",
    11: "Indonesia",
    12: "Japan",
    13: "South Korea",
    14: "Other Asia",
    15: "Total Asia",
    17: "Finland",
    18: "Poland",
    19: "Sweden",
    20: "Total Europe",
    22: "USA",
    23: "Canada",
    24: "Total North America",
    26: "Global"
}

for row_idx, country_name in country_rows.items():
    battery_demand_by_country[country_name] = {}
    for col_idx, year in enumerate(years_nip, start=1):
        battery_demand_by_country[country_name][year] = clean_value(nip_ws.cell_value(row_idx, col_idx))

# End-use percentages (row 35-39, column 1 is 2026)
end_use_pct = {}
end_use_rows = {
    35: "Electric vehicles",
    36: "Energy storage systems",
    37: "Portable electronics",
    38: "Power devices",
    39: "Motive products"
}
for row_idx, category in end_use_rows.items():
    end_use_pct[category] = clean_value(nip_ws.cell_value(row_idx, 1))

# Chemistry percentages (rows 47-62, column 1 is 2026)
chemistry_pct = {}
chemistry_rows = {
    47: "NMC111",
    48: "NMC532",
    49: "NMC622",
    50: "NMC721",
    51: "NMC811",
    52: "NMC High-Ni",
    53: "NCA",
    54: "NMCA",
    55: "NMX",
    56: "LNMO",
    57: "LMRO",
    58: "SSB-NMC811",
    59: "NiMH",
    60: "Total Ni pCAMs",
    61: "Non-Ni pCAMs"
}
for row_idx, chem_name in chemistry_rows.items():
    chemistry_pct[chem_name] = clean_value(nip_ws.cell_value(row_idx, 1))

data["tab_2_ev_battery_nickel"] = {
    "sto_battery_demand_by_country": battery_demand_by_country,
    "sto_end_use_percentages": end_use_pct,
    "sto_chemistry_percentages": chemistry_pct
}

# IHO GlobalTotalAnn - Battery Precursors (row 31)
iho_demand_wb = load_workbook(IHO_DEMAND_FILE)
global_total_ws = iho_demand_wb['GlobalTotalAnn']

iho_battery_demand = {}
# Row 7 has years starting from column 2
for col in range(2, global_total_ws.max_column + 1):
    year = global_total_ws.cell(7, col).value
    if year and isinstance(year, (int, float)):
        year = int(year)
        # Row 31 is Battery precursors
        battery_val = global_total_ws.cell(31, col).value
        if battery_val and not isinstance(battery_val, str):
            iho_battery_demand[year] = clean_value(battery_val)
        else:
            iho_battery_demand[year] = None

data["tab_2_ev_battery_nickel"]["iho_battery_precursors"] = iho_battery_demand

# ============================================================================
# TAB 3: AMBATOVY
# ============================================================================
print("Extracting TAB 3: Ambatovy...")

# IHO NiRefineries sheet
iho_refinery_wb = xlrd.open_workbook(IHO_REFINERY_FILE)
ni_refineries_ws = iho_refinery_wb.sheet_by_name('NiRefineries')

# Row 6 has years starting from column 3
ambatovy_production = {}
years_ref = []
for col in range(3, min(ni_refineries_ws.ncols, 58)):
    year = ni_refineries_ws.cell_value(6, col)
    if year and isinstance(year, (int, float)):
        year = int(year)
        years_ref.append(year)

# Row 8 is Ambatovy label, row 9 has the data
for col_idx, year in enumerate(years_ref, start=3):
    ambatovy_val = ni_refineries_ws.cell_value(9, col_idx)
    ambatovy_production[year] = clean_value(ambatovy_val)

# STO RefinerybyPlant sheet
rfp_ws = sto_wb.sheet_by_name('RefinerybyPlant')

sto_ambatovy_production = {}
# Row 6 has years in columns 2-5
for col in range(2, 6):
    year = rfp_ws.cell_value(6, col)
    if year:
        year = int(year)
        # Row 8 is Ambatovy data
        sto_ambatovy_production[year] = clean_value(rfp_ws.cell_value(8, col))

# Global refinery total for market share calculation
# Row 218 is typically the global total in RefinerybyPlant
global_refinery_total = {}
for col in range(2, 6):
    year = rfp_ws.cell_value(6, col)
    if year:
        year = int(year)
        # Find global total row - scan for it
        for row in range(200, 221):
            if rfp_ws.cell_value(row, 0) and "TOTAL" in str(rfp_ws.cell_value(row, 0)).upper():
                global_refinery_total[year] = clean_value(rfp_ws.cell_value(row, col))
                break

data["tab_3_ambatovy"] = {
    "iho_production": ambatovy_production,
    "sto_production": sto_ambatovy_production,
    "market_share": {}
}

# Calculate market share where data exists
for year in sto_ambatovy_production:
    if year in global_refinery_total and global_refinery_total[year]:
        market_share = safe_divide(sto_ambatovy_production[year], global_refinery_total[year])
        if market_share:
            data["tab_3_ambatovy"]["market_share"][year] = market_share * 100

# ============================================================================
# TAB 4: SOUTH KOREA
# ============================================================================
print("Extracting TAB 4: South Korea...")

# IHO SouthKorea sheet
sk_ws = iho_demand_wb['SouthKorea']

# Extract sector breakdown (rows 10-33 follow similar structure to GlobalTotalAnn)
south_korea_sectors = {}
for row in range(10, 34):
    sector_label = sk_ws.cell(row, 1).value
    if sector_label:
        south_korea_sectors[str(sector_label)] = {}
        for col in range(2, sk_ws.max_column + 1):
            year = sk_ws.cell(7, col).value
            if year and isinstance(year, (int, float)):
                year = int(year)
                value = sk_ws.cell(row, col).value
                if value and not isinstance(value, str):
                    south_korea_sectors[str(sector_label)][year] = clean_value(value)

# IHO South Korea SS Cap sheet
sk_ss_cap_ws = iho_demand_wb['South Korea SS Cap']

plants_capacity = {}
# Rows 8-10 have plant names in column 2
plant_rows = {
    8: "Bae Myung Metal",
    9: "Posco",
    10: "Seah Besteel"
}

for row_idx, plant_name in plant_rows.items():
    plants_capacity[plant_name] = {}
    for col in range(3, sk_ss_cap_ws.max_column + 1):
        year = sk_ss_cap_ws.cell(7, col).value
        if year and isinstance(year, (int, float)):
            year = int(year)
            capacity = sk_ss_cap_ws.cell(row_idx, col).value
            if capacity and isinstance(capacity, (int, float)):
                plants_capacity[plant_name][year] = capacity

# STO NiInPrecursors - Row 13 is South Korea battery demand
sk_battery_demand = {}
for col_idx, year in enumerate(years_nip, start=1):
    sk_battery_demand[year] = clean_value(nip_ws.cell_value(13, col_idx))

data["tab_4_south_korea"] = {
    "iho_sector_breakdown": south_korea_sectors,
    "iho_ss_plant_capacities": plants_capacity,
    "sto_battery_demand": sk_battery_demand
}

# ============================================================================
# TAB 5: SUPPLY
# ============================================================================
print("Extracting TAB 5: Supply...")

# STO MineProduction sheet
mine_prod_ws = sto_wb.sheet_by_name('MineProduction')
mine_production = {}
# Row 6 has years starting at col 1, Row 52 is "Forecast mine production"
for col in range(1, min(8, mine_prod_ws.ncols)):
    year = mine_prod_ws.cell_value(6, col)
    if year and isinstance(year, (int, float)):
        year = int(year)
        mine_production[year] = clean_value(mine_prod_ws.cell_value(52, col))

# STO SmelterProduction sheet
smelter_prod_ws = sto_wb.sheet_by_name('SmelterProduction')
smelter_production = {}
# Row 6 has years starting at col 1, Row 50 is "Total World"
for col in range(1, min(8, smelter_prod_ws.ncols)):
    year = smelter_prod_ws.cell_value(6, col)
    if year and isinstance(year, (int, float)):
        year = int(year)
        smelter_production[year] = clean_value(smelter_prod_ws.cell_value(50, col))

# STO RefineryProduction sheet
refinery_prod_ws = sto_wb.sheet_by_name('RefineryProduction')
refinery_production = {}
# Row 6 has years starting at col 1, Row 52 is "Forecast refinery production"
for col in range(1, min(8, refinery_prod_ws.ncols)):
    year = refinery_prod_ws.cell_value(6, col)
    if year and isinstance(year, (int, float)):
        year = int(year)
        refinery_production[year] = clean_value(refinery_prod_ws.cell_value(52, col))

# STO GlobalBalance sheet
gb_ws = sto_wb.sheet_by_name('GlobalBalance')
global_balance = {}
# Structure varies; typically rows 6-7 have years and key metrics
for col in range(2, min(6, gb_ws.ncols)):
    year = gb_ws.cell_value(6, col)
    if year:
        year = int(year)
        global_balance[year] = {
            "supply_kt": clean_value(gb_ws.cell_value(8, col)) if gb_ws.nrows > 8 else None,
            "consumption_kt": clean_value(gb_ws.cell_value(9, col)) if gb_ws.nrows > 9 else None,
            "balance_kt": clean_value(gb_ws.cell_value(10, col)) if gb_ws.nrows > 10 else None
        }

data["tab_5_supply"] = {
    "sto_mine_production": mine_production,
    "sto_smelter_production": smelter_production,
    "sto_refinery_production": refinery_production,
    "sto_global_balance": global_balance
}

# ============================================================================
# Save JSON
# ============================================================================
output_path = str(BASE_DIR / "data/woodmac_data.json")
with open(output_path, 'w') as f:
    json.dump(data, f, indent=2)

print(f"\nData extraction complete!")
print(f"Output saved to: {output_path}")
print(f"Total size: {len(json.dumps(data))} bytes")
